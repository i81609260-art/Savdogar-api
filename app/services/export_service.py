"""Generic Excel/PDF table export helpers.

Two small, reusable exporters that turn a header row + data rows into
downloadable bytes. Routers gather the data and call these — no per-report
special casing here.
"""

from io import BytesIO
from typing import Any, Sequence

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def rows_to_excel(
    title: str, headers: Sequence[str], rows: Sequence[Sequence[Any]]
) -> bytes:
    """Build an .xlsx workbook from a header row and data rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = (title[:31] or "Hisobot")

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")  # indigo-600
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r, row in enumerate(rows, start=2):
        for col, value in enumerate(row, start=1):
            ws.cell(row=r, column=col, value=value)

    # Auto-size columns to their widest cell (capped).
    for col in range(1, len(headers) + 1):
        width = len(str(headers[col - 1]))
        for row in rows:
            if col - 1 < len(row):
                width = max(width, len(str(row[col - 1])))
        ws.column_dimensions[get_column_letter(col)].width = min(width + 4, 50)

    ws.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _latin1(value: Any) -> str:
    """fpdf2 core fonts are Latin-1 only; downgrade Uzbek/Unicode safely."""
    if value is None:
        return ""
    text = str(value)
    for src in ("ʻ", "ʼ", "‘", "’", "`"):
        text = text.replace(src, "'")
    text = text.replace("“", '"').replace("”", '"').replace("–", "-")
    return text.encode("latin-1", "replace").decode("latin-1")


class _TablePDF(FPDF):
    """Landscape A4 table with a branded header/footer."""

    doc_title = "Hisobot"

    def header(self) -> None:
        self.set_font("Helvetica", "B", 14)
        self.set_fill_color(79, 70, 229)
        self.rect(0, 0, self.w, 18, "F")
        self.set_text_color(255, 255, 255)
        self.set_xy(0, 4)
        self.cell(0, 10, _latin1(self.doc_title), align="C")
        self.set_text_color(0, 0, 0)
        self.set_y(24)

    def footer(self) -> None:
        self.set_y(-12)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(120, 120, 120)
        self.cell(0, 8, f"savdogar.uz  |  {self.page_no()}", align="C")


def rows_to_pdf(
    title: str, headers: Sequence[str], rows: Sequence[Sequence[Any]]
) -> bytes:
    """Build a landscape PDF table from a header row and data rows."""
    pdf = _TablePDF(orientation="L", unit="mm", format="A4")
    pdf.doc_title = title
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    usable = pdf.w - 2 * pdf.l_margin
    col_w = usable / max(len(headers), 1)

    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(230, 230, 245)
    for name in headers:
        pdf.cell(col_w, 8, _latin1(name), border=1, fill=True, align="C")
    pdf.ln()

    pdf.set_font("Helvetica", "", 8)
    fill = False
    for row in rows:
        for i in range(len(headers)):
            value = row[i] if i < len(row) else ""
            text = _latin1(value)
            if len(text) > 42:
                text = text[:39] + "..."
            pdf.cell(col_w, 7, text, border=1, fill=fill, align="L")
        pdf.ln()
        fill = not fill

    return bytes(pdf.output())
