"""Price-list tahlili — operator yuborgan narxlar ro'yxatini strukturaga solish.

Nima uchun bu birinchi
----------------------
O'zbekiston bozorida operatorlar narxni har kuni Telegram kanal va agent
guruhlariga tashlaydi: Excel jadval, PDF, ba'zan oddiy xabar. Turagent
ularni qo'lda o'qib chiqadi. Bu — bozorning haqiqiy narx oqimi, va uni
o'qish uchun operator sayti ham, login-parol ham, uning roziligi ham
kerak emas.

Nima ishlatiladi
----------------
Tashqi API **yo'q**. Ustun nomlarini tanish — taxalluslar ro'yxati,
yo'nalish/ovqat/yulduz — `tour_taxonomy`, narx — regex. Ya'ni Tella bilan
bir xil yondashuv: qoida bilan yechiladigan narsaga model chaqirilmaydi.

Qo'llab-quvvatlanadigan formatlar:
  * `.xlsx` / `.xlsm` — openpyxl (allaqachon o'rnatilgan)
  * `.csv`            — standart kutubxona
  * oddiy matn        — Telegram xabari, satrma-satr
  * `.pdf`            — matn ajratib olinadi (pypdf bo'lsa)

Rasm (jpg/png) — hozircha yo'q. U yagona format bo'lib, mahalliy vositalar
bilan ishonchli o'qilmaydi; keyinroq alohida hal qilinadi.
"""

from __future__ import annotations

import csv
import io
import logging
import re
from dataclasses import dataclass
from typing import Any, Iterable, Optional

from app.services.operator_connector import RawOffer
from app.services.tour_taxonomy import (
    CURRENCY_ALIASES,
    Destination,
    match_board,
    match_currency,
    match_destinations,
    match_star,
    normalize,
)

log = logging.getLogger(__name__)

MAX_ROWS = 5_000
HEADER_SCAN_ROWS = 15


# --------------------------------------------------------------------------
# Ustun nomlari
# --------------------------------------------------------------------------
COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "hotel_name": (
        "otel", "mehmonxona", "gostinitsa", "отель", "гостиница", "hotel",
        "hotel name", "nomi", "название", "naimenovanie", "наименование",
    ),
    "price_gross": (
        "narx", "narxi", "цена", "стоимость", "price", "sotuv narxi",
        "brutto", "брутто", "prays", "sum", "summa", "сумма", "tsena",
    ),
    "price_net": ("netto", "нетто", "net", "net narx", "нетто цена", "sof narx"),
    "commission_pct": ("komissiya", "комиссия", "commission", "komis"),
    "board": ("ovqat", "ovqatlanish", "питание", "board", "meal", "pitanie"),
    "star": ("yulduz", "звезд", "звёзд", "star", "stars", "kategoriya", "категория"),
    "nights": (
        "kecha", "kun", "ночей", "ночи", "nights", "noch", "muddat",
        "длительность", "продолжительность",
    ),
    "city": ("shahar", "kurort", "город", "курорт", "city", "resort", "yonalish",
             "направление", "napravlenie"),
    "country": ("davlat", "mamlakat", "страна", "country"),
    "date_from": (
        "sana", "chiqish", "дата", "дата вылета", "вылет", "date", "zayezd",
        "заезд", "ketish",
    ),
    "room": ("xona", "номер", "room", "tip nomera", "тип номера"),
    "currency": ("valyuta", "валюта", "currency"),
}


def _match_column(header: Any) -> Optional[str]:
    """Ustun sarlavhasini kanonik maydonga bog'laydi."""
    text = normalize(str(header or ""))
    if not text:
        return None
    # Aniqroq (uzunroq) taxallus ustun turadi: "netto narx" -> price_net,
    # "narx" -> price_gross. Aks holda "netto narx" ichidagi "narx" g'olib
    # bo'lib, netto narx sotuv narxi deb yozilardi.
    best: tuple[int, Optional[str]] = (0, None)
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            alias_n = normalize(alias)
            if alias_n and alias_n in text and len(alias_n) > best[0]:
                best = (len(alias_n), field)
    return best[1]


def map_columns(headers: Iterable[Any]) -> dict[int, str]:
    """Sarlavha qatoridan {ustun raqami: maydon} yasaydi."""
    mapping: dict[int, str] = {}
    used: set[str] = set()
    for index, header in enumerate(headers):
        field = _match_column(header)
        # Bir maydon ikki ustunga bog'lanmasin — birinchisi qoladi.
        if field and field not in used:
            mapping[index] = field
            used.add(field)
    return mapping


# --------------------------------------------------------------------------
# Narx
# --------------------------------------------------------------------------
# Valyuta taxalluslari BITTA manbadan — `CURRENCY_ALIASES`.
#
# Ilgari bu ro'yxat shu faylda qo'lda takrorlangan edi va vaqt o'tib
# taksonomiyadan ajralib ketdi: u yerda "dollar", "сўм", "долл", "rubl",
# "euro" bor edi, bu yerdagi nusxada yo'q. Natijada "890 dollar" narxi
# tanilmasdi. Endi ro'yxat bitta joyda — ajralib ketishi mumkin emas.
_CURRENCY_WORDS = "|".join(
    sorted(
        (
            # Apostrof turlicha yoziladi: so'm / soʻm / so'm / som
            re.escape(alias).replace("'", "['‘’ʻ]?")
            for aliases in CURRENCY_ALIASES.values()
            for alias in aliases
        ),
        key=len,
        reverse=True,  # "доллар" "долл" dan oldin tekshirilsin
    )
)

# `(?!\w)` — "500 summa" ichidagi "sum" ni valyuta deb o'qimasin.
_CURRENCY_IN_PRICE = re.compile(rf"({_CURRENCY_WORDS})(?!\w)", re.I)

# Mingliklar ajratgichi bo'la oladigan bo'sh joy.
#
# TAB VA YANGI QATOR ATAYLAB KIRMAYDI: jadvalda ular USTUNLARNI
# ajratadi. `\s` ishlatilganda brauzerdan olingan "7<TAB>890 USD"
# (7 kecha, 890 dollar) bitta son — 7890 — bo'lib o'qilardi va agent
# narxni to'qqiz barobar oshirib qo'yardi.
_SEP_CHARS = " \u00a0\u202f"
_SEP = rf"[{_SEP_CHARS}]"


def parse_price(value: Any) -> tuple[Optional[float], Optional[str]]:
    """Narxni va valyutani ajratadi.

    Ajratgichlar mintaqaga qarab teskari ma'no beradi: `1.200` yevropada
    ming sakkiz yuz emas — ming ikki yuz, amerikada esa bir butun ikki.
    Qoida: agar ikkala belgi ham bo'lsa, OXIRGISI kasr ajratgichi. Bitta
    belgi bo'lsa va undan keyin roppa-rosa 3 raqam kelsa — minglik.
    """
    if value is None:
        return None, None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return (float(value) or None), None

    text = str(value).strip()
    if not text:
        return None, None

    currency = match_currency(text)
    if not currency:
        found = _CURRENCY_IN_PRICE.search(text)
        if found:
            currency = match_currency(found.group(1))

    # Tab yoki yangi qator — USTUN chegarasi. Ular qolsa qo'shni
    # ustundagi son narxga yopishib ketardi ("7<TAB>890" -> 7890).
    # Valyuta qaysi bo'lakda bo'lsa, narx ham o'sha yerda.
    if "\t" in text or "\n" in text:
        boklar = [b for b in re.split(r"[\t\n]+", text) if re.search(r"\d", b)]
        if boklar:
            valyutali = [b for b in boklar if _CURRENCY_IN_PRICE.search(b)]
            text = (valyutali or boklar)[-1].strip()

    # Faqat raqam va ajratgichlar qoldiriladi.
    cleaned = re.sub(rf"[^\d.,{_SEP_CHARS}]", "", text).strip()
    cleaned = re.sub(rf"{_SEP}+", "", cleaned)  # bo'sh joy — doim minglik
    if not cleaned or not re.search(r"\d", cleaned):
        return None, currency

    has_dot, has_comma = "." in cleaned, "," in cleaned
    if has_dot and has_comma:
        decimal_sep = "." if cleaned.rfind(".") > cleaned.rfind(",") else ","
        thousands_sep = "," if decimal_sep == "." else "."
        cleaned = cleaned.replace(thousands_sep, "").replace(decimal_sep, ".")
    elif has_dot or has_comma:
        sep = "." if has_dot else ","
        tail = cleaned.rsplit(sep, 1)[1]
        if len(tail) == 3 and cleaned.count(sep) >= 1:
            cleaned = cleaned.replace(sep, "")   # minglik
        else:
            cleaned = cleaned.replace(sep, ".")  # kasr

    try:
        amount = float(cleaned)
    except ValueError:
        return None, currency

    # "5 mln", "800 ming" — so'mdagi narxlar deyarli doim shunday yoziladi.
    # Ko'paytuvchisiz 5 000 000 o'rniga 5 yozilib qolardi.
    amount *= _multiplier(text)
    return (amount if amount > 0 else None), currency


_MULTIPLIERS: tuple[tuple[str, int], ...] = (
    ("mln", 1_000_000), ("million", 1_000_000), ("млн", 1_000_000),
    ("ming", 1_000), ("тыс", 1_000),
)


def _multiplier(text: str) -> int:
    lowered = text.lower()
    for word, factor in _MULTIPLIERS:
        if word in lowered:
            return factor
    return 1


def _parse_int(value: Any, low: int, high: int) -> Optional[int]:
    if value is None:
        return None
    match = re.search(r"\d{1,3}", str(value))
    if not match:
        return None
    number = int(match.group())
    return number if low <= number <= high else None


# --------------------------------------------------------------------------
# Natija
# --------------------------------------------------------------------------
@dataclass
class PricelistResult:
    """Tahlil natijasi va diagnostikasi.

    `skipped` va `warnings` ataylab qaytariladi: agent price-list yuklaganda
    "142 tadan 8 tasi o'qilmadi" deb ko'rsatish kerak. Jim yutilgan qatorlar
    — yo'qolgan narx demakdir.
    """

    offers: list[RawOffer]
    total_rows: int = 0
    skipped: int = 0
    warnings: list[str] = None  # type: ignore[assignment]

    def __post_init__(self) -> None:
        if self.warnings is None:
            self.warnings = []


# --------------------------------------------------------------------------
# Jadval (Excel / CSV)
# --------------------------------------------------------------------------
def _rows_from_xlsx(content: bytes) -> list[list[Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    rows: list[list[Any]] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(values_only=True):
            rows.append(list(row))
            if len(rows) >= MAX_ROWS:
                return rows
    return rows


def _rows_from_csv(content: bytes) -> list[list[Any]]:
    text = content.decode("utf-8", errors="replace")
    # Ajratgich vergul yoki nuqta-vergul bo'lishi mumkin (rus Excel'i `;`).
    sample = text[:4000]
    delimiter = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    return [row for _, row in zip(range(MAX_ROWS), reader)]


def _find_header(rows: list[list[Any]]) -> tuple[int, dict[int, str]]:
    """Sarlavha qatorini topadi — eng ko'p ustuni tanilgan qator.

    Price-list'lar deyarli hech qachon birinchi qatordan boshlanmaydi:
    yuqorida logotip, sarlavha, sana bo'ladi.
    """
    best_index, best_map = -1, {}
    for index, row in enumerate(rows[:HEADER_SCAN_ROWS]):
        mapping = map_columns(row)
        if len(mapping) > len(best_map):
            best_index, best_map = index, mapping
    # Mehmonxona va narx bo'lmasa bu sarlavha emas.
    if not {"hotel_name", "price_gross", "price_net"} & set(best_map.values()):
        return -1, {}
    return best_index, best_map


def parse_table(content: bytes, filename: str = "") -> PricelistResult:
    """Excel yoki CSV price-list'ni o'qiydi."""
    name = (filename or "").lower()
    try:
        rows = _rows_from_xlsx(content) if name.endswith((".xlsx", ".xlsm")) \
            else _rows_from_csv(content)
    except Exception as exc:  # noqa: BLE001
        return PricelistResult(offers=[], warnings=[f"Fayl o'qilmadi: {exc}"])

    header_index, mapping = _find_header(rows)
    if header_index < 0:
        return PricelistResult(
            offers=[], total_rows=len(rows),
            warnings=["Sarlavha qatori topilmadi (mehmonxona/narx ustuni yo'q)"],
        )

    offers: list[RawOffer] = []
    skipped = 0
    for row in rows[header_index + 1:]:
        offer = _row_to_offer(row, mapping)
        if offer is None:
            if any(str(c or "").strip() for c in row):
                skipped += 1
            continue
        offers.append(offer)

    warnings = []
    if skipped:
        warnings.append(f"{skipped} qator o'qilmadi (mehmonxona yoki narx yo'q)")
    return PricelistResult(
        offers=offers, total_rows=len(rows) - header_index - 1,
        skipped=skipped, warnings=warnings,
    )


def _row_to_offer(row: list[Any], mapping: dict[int, str]) -> Optional[RawOffer]:
    """Bitta qatordan taklif yasaydi. Yaroqsiz bo'lsa `None`."""
    values: dict[str, Any] = {}
    for index, field in mapping.items():
        if index < len(row):
            values[field] = row[index]

    hotel = str(values.get("hotel_name") or "").strip()
    price_gross, currency_g = parse_price(values.get("price_gross"))
    price_net, currency_n = parse_price(values.get("price_net"))
    if not hotel or (price_gross is None and price_net is None):
        return None

    row_text = " ".join(str(v) for v in row if v is not None)
    destinations = match_destinations(str(values.get("city") or "")
                                      or str(values.get("country") or "") or row_text)

    return RawOffer(
        hotel_name=hotel[:300],
        price_gross=price_gross,
        price_net=price_net,
        currency=(
            str(values.get("currency") or "").strip().upper() or None
        ) or currency_g or currency_n or match_currency(row_text),
        board=(match_board(str(values.get("board") or "")) or match_board(row_text)),
        star=(match_star(str(values.get("star") or "")) or match_star(row_text)),
        room=str(values.get("room") or "").strip()[:120] or None,
        city=_first_name(destinations, country=False),
        country=_first_name(destinations, country=True),
        nights=_parse_int(values.get("nights"), 1, 60),
        commission_pct=_parse_int(values.get("commission_pct"), 0, 100),
        raw={"row": [str(c) if c is not None else None for c in row]},
    )


def _first_name(destinations: list[Destination], country: bool) -> Optional[str]:
    for dest in destinations:
        if dest.is_country == country:
            return dest.name_uz
    return None


# --------------------------------------------------------------------------
# Matn (Telegram xabari)
# --------------------------------------------------------------------------
# "Rixos Downtown 5* UAI — $850" ko'rinishidagi satrlar.
#
# Narxi tanilmagan satr sarlavha (kontekst) o'rnida qabul qilinadi va jimgina
# tashlab yuboriladi — shuning uchun bu yerdagi valyuta ro'yxati to'liq
# bo'lishi muhim. U `_CURRENCY_WORDS` orqali taksonomiyaga bog'langan.
_PRICE_IN_LINE = re.compile(
    # Belgi narxdan oldin: "$850". Faqat belgilar — "usd 850" deb yozilmaydi.
    rf"(?:\$|€|₽){_SEP}*\d[\d.,{_SEP_CHARS}]*|"
    # Valyuta narxdan keyin: "850 $", "890 dollar", "12 000 000 so'm".
    # `(?!\w)` — "500 summa" ichidagi "sum" ni valyuta deb o'qimasin.
    rf"\d[\d.,{_SEP_CHARS}]*{_SEP}*(?:{_CURRENCY_WORDS})(?!\w)",
    re.I,
)


def parse_text(text: str) -> PricelistResult:
    """Telegram xabari yoki oddiy matnli price-list.

    Sarlavha satrlari (yo'nalish, kecha) **kontekst** bo'lib, undan keyingi
    narxli satrlarga tarqaladi. Odatdagi post shunday tuziladi:

        ANTALYA 7 kecha
        Rixos Downtown 5* UAI — $850
        Delphin Imperial 5* AI — $720
    """
    lines = [ln.strip() for ln in (text or "").splitlines()]
    offers: list[RawOffer] = []

    ctx_destinations: list[Destination] = []
    ctx_nights: Optional[int] = None
    skipped = 0

    for line in lines:
        if not line:
            continue

        price_match = _PRICE_IN_LINE.search(line)
        if not price_match:
            # Kontekst satri — yo'nalish yoki muddatni yangilaydi.
            found = match_destinations(line)
            if found:
                ctx_destinations = found
            nights = _nights_in(line)
            if nights:
                ctx_nights = nights
            continue

        amount, currency = parse_price(price_match.group())
        if amount is None:
            skipped += 1
            continue

        # Mehmonxona nomi — narx va belgilardan tozalangan qism.
        name = line[: price_match.start()]

        # Ustunli satrda (brauzerdan olingan jadval) nom BIRINCHI
        # katakda. Qolgan kataklar — yulduz, ovqat, kecha — nomga
        # yopishib "Rixos Downtown Antalya\t\t\t7" bo'lib chiqardi.
        # Ayni paytda o'sha kataklardan kecha sonini ham olamiz.
        if "\t" in name:
            kataklar = [k.strip() for k in name.split("\t") if k.strip()]
            if kataklar:
                for katak in kataklar[1:]:
                    kecha = _nights_in(katak) or (
                        int(katak) if katak.isdigit() and 1 <= int(katak) <= 60
                        else None
                    )
                    if kecha:
                        ctx_nights = kecha
                        break
                name = kataklar[0]

        name = re.sub(r"[—–\-:|•·]+\s*$", "", name).strip()
        name = re.sub(r"\b\d\s*\*|\b(UAI|AI|HB\+?|FB\+?|BB|RO)\b", "", name,
                      flags=re.I).strip(" -—–:|")
        if len(name) < 3:
            skipped += 1
            continue

        line_destinations = match_destinations(line) or ctx_destinations
        offers.append(RawOffer(
            hotel_name=name[:300],
            price_gross=amount,
            currency=currency or match_currency(line),
            board=match_board(line),
            star=match_star(line),
            nights=_nights_in(line) or ctx_nights,
            city=_first_name(line_destinations, country=False),
            country=_first_name(line_destinations, country=True),
            raw={"line": line},
        ))

    warnings = [f"{skipped} satr o'qilmadi"] if skipped else []
    return PricelistResult(
        offers=offers, total_rows=len([ln for ln in lines if ln]),
        skipped=skipped, warnings=warnings,
    )


def _nights_in(line: str) -> Optional[int]:
    match = re.search(
        r"(\d{1,2})\s*(?:kecha|kun|ночей|ночи|ноч|nights?)", normalize(line)
    )
    if not match:
        return None
    value = int(match.group(1))
    return value if 1 <= value <= 60 else None


# --------------------------------------------------------------------------
# Umumiy kirish nuqtasi
# --------------------------------------------------------------------------
def parse_pricelist(
    content: bytes | str, filename: str = "", content_type: str = ""
) -> PricelistResult:
    """Formatni aniqlab, mos tahlilchini chaqiradi."""
    if isinstance(content, str):
        return parse_text(content)

    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm", ".csv")) or "spreadsheet" in content_type:
        return parse_table(content, filename)

    if name.endswith(".pdf") or "pdf" in content_type:
        text = _pdf_text(content)
        if text is None:
            return PricelistResult(
                offers=[],
                warnings=["PDF o'qish uchun `pypdf` o'rnatilmagan"],
            )
        return parse_text(text)

    if _is_image(name, content_type):
        # Rasm — yagona format bo'lib, sinxron tahlil qilib bo'lmaydi
        # (tashqi xizmatga so'rov ketadi). Shuning uchun `parse_pricelist`
        # uni bajarmaydi; chaqiruvchi `parse_pricelist_async` ni ishlatsin.
        return PricelistResult(
            offers=[],
            warnings=["Rasmli price-list uchun `parse_pricelist_async` kerak"],
        )

    return parse_text(content.decode("utf-8", errors="replace"))


IMAGE_EXT = (".jpg", ".jpeg", ".png", ".webp", ".heic")


def _is_image(filename: str, content_type: str) -> bool:
    return filename.endswith(IMAGE_EXT) or "image" in (content_type or "")


async def parse_pricelist_async(
    content: bytes | str, filename: str = "", content_type: str = ""
) -> PricelistResult:
    """`parse_pricelist` bilan bir xil, lekin rasmni ham qo'llab-quvvatlaydi.

    Rasm tahlili tashqi xizmatga so'rov yuboradi, shuning uchun alohida
    asinxron kirish nuqtasi. Qolgan formatlar uchun farq yo'q — o'sha
    sinxron tahlilchi chaqiriladi.
    """
    name = (filename or "").lower()
    if isinstance(content, bytes) and _is_image(name, content_type):
        from app.services.pricelist_image import parse_image

        return await parse_image(content, content_type or "image/jpeg")
    return parse_pricelist(content, filename, content_type)


def _pdf_text(content: bytes) -> Optional[str]:
    """PDF dan matn ajratadi. `pypdf` bo'lmasa `None`."""
    try:
        from pypdf import PdfReader
    except ImportError:
        return None
    try:
        reader = PdfReader(io.BytesIO(content))
        return "\n".join((page.extract_text() or "") for page in reader.pages)
    except Exception as exc:  # noqa: BLE001
        log.warning("PDF o'qilmadi: %s", exc)
        return None
